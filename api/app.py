from flask import Flask, request, send_file, render_template
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import os
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder="../templates")
UPLOAD_FOLDER = "/tmp/uploads"
OUTPUT_FOLDER = "/tmp/outputs"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "bmp"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def get_scaled_dimensions(width, height, max_dimension=100):
    if width <= max_dimension and height <= max_dimension:
        return width, height
    ratio = min(max_dimension / width, max_dimension / height)
    return int(width * ratio), int(height * ratio)

def image_to_excel(input_image_path, output_excel_path, grid_size="free"):
    try:
        img = Image.open(input_image_path).convert("RGB")
        width, height = img.size
        if grid_size == "free":
            target_width, target_height = get_scaled_dimensions(width, height)
            img = img.resize((target_width, target_height), Image.LANCZOS)
        else:
            target_width = target_height = int(grid_size)
            img = img.resize((target_width, target_height), Image.LANCZOS)
        img_array = np.array(img)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ImageArt"
        for col in range(1, target_width + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2
        for row in range(1, target_height + 1):
            ws.row_dimensions[row].height = 12
        for row in range(target_height):
            for col in range(target_width):
                r, g, b = img_array[row, col]
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell = ws.cell(row=row + 1, column=col + 1)
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        wb.save(output_excel_path)
        return True, "Success", target_width, target_height
    except Exception as e:
        return False, str(e), None, None

@app.route("/", methods=["GET", "POST"])
def upload_image():
    message = None
    error = False
    if request.method == "POST":
        if "image" not in request.files:
            message = "No file uploaded"
            error = True
        else:
            file = request.files["image"]
            if file.filename == "":
                message = "No file selected"
                error = True
            elif not allowed_file(file.filename):
                message = "Invalid file type. Use PNG, JPG, JPEG, or BMP"
                error = True
            else:
                try:
                    grid_size = request.form.get("grid_size", "100")
                    if grid_size not in ["25", "50", "100", "free"]:
                        message = "Invalid grid size. Choose 25x25, 50x50, 100x100, or Free Size"
                        error = True
                    else:
                        filename = secure_filename(file.filename)
                        input_path = os.path.join(UPLOAD_FOLDER, filename)
                        output_path = os.path.join(OUTPUT_FOLDER, f"{os.path.splitext(filename)[0]}_art.xlsx")
                        file.save(input_path)
                        success, result, width, height = image_to_excel(input_path, output_path, grid_size)
                        if success:
                            size_str = f"{width}x{height}" if grid_size == "free" else f"{grid_size}x{grid_size}"
                            return send_file(
                                output_path,
                                as_attachment=True,
                                download_name=f"image_art_{size_str}.xlsx"
                            )
                        else:
                            message = f"Error processing image: {result}"
                            error = True
                finally:
                    if "input_path" in locals() and os.path.exists(input_path):
                        os.remove(input_path)
                    if "output_path" in locals() and os.path.exists(output_path):
                        os.remove(output_path)
    return render_template("index.html", message=message, error=error)
